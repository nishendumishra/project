"""
xlsx_ordered_extractor_no_dup_fixed.py

Improved extractor that preserves order, avoids duplicate outputs in JSON/MD,
and places images exactly where they occur. Mark cells as visited so text
isn't repeatedly detected as table cells. Final deduplication prevents
accidental repeats in output.
"""

import os
import json
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
from io import BytesIO

# ---------- CONFIG ----------
OUTPUT_DIR = "output"
IMAGE_DIR = os.path.join(OUTPUT_DIR, "images")
os.makedirs(IMAGE_DIR, exist_ok=True)
# ----------------------------

def get_image_description(image_path):
    """
    Placeholder: replace with Ollama or your HTTP LLaMA Vision API call.
    Keep it robust to exceptions and return a string.
    """
    try:
        # Example simple caption; replace with real API call
        return f"Description for {os.path.basename(image_path)}"
    except Exception as e:
        return f"Error describing image: {e}"

def expand_merged_cells(ws):
    """
    Return a dict mapping (r,c) 0-based -> value and the actual grid size (rows,cols).
    Propagates merged cell top-left values to all cells in the merged range.
    """
    max_row = ws.max_row
    max_col = ws.max_column
    cell_map = {}
    # Fill base values
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell_map[(r-1, c-1)] = ws.cell(row=r, column=c).value

    # Propagate merged ranges
    for merged in ws.merged_cells.ranges:
        min_row, min_col, maxr, maxc = merged.min_row, merged.min_col, merged.max_row, merged.max_col
        top_val = ws.cell(row=min_row, column=min_col).value
        for rr in range(min_row - 1, maxr):
            for cc in range(min_col - 1, maxc):
                cell_map[(rr, cc)] = top_val

    return cell_map, max_row, max_col

def image_anchor_pos(image):
    """
    Try to get 0-based (row, col) anchor for an openpyxl image.
    If not available, returns None.
    """
    anchor = getattr(image, "anchor", None)
    if not anchor:
        return None
    # try multiple possible attribute names robustly
    frm = getattr(anchor, "_from", None) or getattr(anchor, "from_", None) or getattr(anchor, "from", None)
    if frm:
        row = getattr(frm, "row", None)
        col = getattr(frm, "col", None)
        if row is not None and col is not None:
            return int(row), int(col)
    # fallback: try 'row'/'col' directly
    row = getattr(anchor, "row", None)
    col = getattr(anchor, "col", None)
    if row is not None and col is not None:
        return int(row), int(col)
    return None

def save_image_from_openpyxl(image_obj, sheetname, idx):
    """
    Save image bytes from openpyxl image object to IMAGE_DIR and return path.
    """
    img_path = os.path.join(IMAGE_DIR, f"{sheetname}_image_{idx+1}.png")
    try:
        data = None
        # Many openpyxl Image objects have ._data() that returns bytes
        if hasattr(image_obj, "_data"):
            data = image_obj._data()
        # Some have .image - attempt to access .image if it's a PIL image
        if data is None and hasattr(image_obj, "image"):
            inner = image_obj.image
            try:
                if hasattr(inner, "fp"):
                    inner.fp.seek(0)
                    im = PILImage.open(inner.fp)
                    im.save(img_path)
                    return img_path
                elif isinstance(inner, PILImage.Image):
                    inner.save(img_path)
                    return img_path
            except Exception:
                pass
        if data:
            im = PILImage.open(BytesIO(data))
            im.save(img_path)
            return img_path
    except Exception:
        pass

    # Fallback placeholder (so pipeline does not crash)
    placeholder = PILImage.new("RGB", (200, 80), color=(230, 230, 230))
    placeholder.save(img_path)
    return img_path

def detect_table_block(grid, visited, start_r, max_row, max_col):
    """
    Detect a table starting at row start_r.
    - A table must have at least two consecutive rows where each row has >=2 unvisited non-empty cells.
    - Returns (sr, er, sc, ec) as 0-based inclusive bbox, or None.
    """
    candidate_rows = []
    r = start_r
    # look ahead up to 50 rows (safe limit)
    while r < max_row and len(candidate_rows) < 50:
        nonempty_cols = [c for c in range(max_col)
                         if (r, c) not in visited and grid.get((r, c)) is not None and str(grid.get((r,c))).strip() != ""]
        if len(nonempty_cols) >= 2:
            candidate_rows.append(nonempty_cols)
            r += 1
            # require at least 2 rows of this kind
            if len(candidate_rows) >= 2:
                break
        else:
            break

    if len(candidate_rows) < 2:
        return None

    # Determine col span using union of indices in candidate_rows
    all_indices = sorted({idx for sub in candidate_rows for idx in sub})
    if not all_indices:
        return None
    sc = all_indices[0]
    ec = all_indices[-1]
    sr = start_r
    er = sr + len(candidate_rows) - 1
    # Expand er downward to include further rows that have at least one unvisited non-empty in the same span
    rr = er + 1
    while rr < max_row:
        row_has = False
        for c in range(sc, ec + 1):
            if (rr, c) not in visited and grid.get((rr, c)) is not None and str(grid.get((rr,c))).strip() != "":
                row_has = True
                break
        if row_has:
            rr += 1
        else:
            break
    er = rr - 1 if rr > er + 1 else er
    return (sr, er, sc, ec)

def rows_to_table_dicts(grid, sr, er, sc, ec, visited):
    """
    Convert the table bbox to headers + list of row dicts. Marks visited.
    Returns (headers, rows)
    """
    # header = row sr
    headers = []
    for c in range(sc, ec + 1):
        h = grid.get((sr, c))
        if h is None or str(h).strip() == "":
            headers.append(f"Column_{c - sc + 1}")
        else:
            headers.append(str(h))
        visited.add((sr, c))  # mark header visited

    rows = []
    for rr in range(sr + 1, er + 1):
        row_dict = {}
        any_nonempty = False
        for i, c in enumerate(range(sc, ec + 1)):
            val = grid.get((rr, c))
            if val is not None and str(val).strip() != "":
                any_nonempty = True
                row_dict[headers[i]] = val
            else:
                row_dict[headers[i]] = ""
            visited.add((rr, c))
        if any_nonempty:
            rows.append(row_dict)
    return headers, rows

def extract_ordered_sheet(ws, get_image_desc_fn):
    """
    Returns ordered, deduplicated list of items for the worksheet.
    Each item: dict with type in {text, table, image} and row/col (0-based).
    """
    grid, max_row, max_col = expand_merged_cells(ws)
    visited = set()
    items = []

    # Preprocess images: save and compute anchor position
    images_info = []
    for i, img in enumerate(ws._images):
        if not isinstance(img, XLImage):
            continue
        pos = image_anchor_pos(img)
        img_path = save_image_from_openpyxl(img, ws.title, i)
        desc = get_image_desc_fn(img_path)
        if pos is None:
            # place after last row
            images_info.append({"row": max_row, "col": 0, "path": img_path, "desc": desc})
        else:
            images_info.append({"row": pos[0], "col": pos[1], "path": img_path, "desc": desc})

    # Build lookup for images anchored at row
    images_by_row = {}
    for info in images_info:
        images_by_row.setdefault(info["row"], []).append(info)

    r = 0
    while r < max_row:
        # Insert images anchored at this row first (in column order)
        if r in images_by_row:
            for imginfo in sorted(images_by_row[r], key=lambda x: x["col"]):
                # avoid duplicates if same path already appended
                items.append({
                    "type": "image",
                    "path": imginfo["path"],
                    "description": imginfo["desc"],
                    "row": imginfo["row"],
                    "col": imginfo["col"]
                })

        # Find unvisited non-empty cells in this row
        row_nonvisited_nonempty = [c for c in range(max_col)
                                   if (r, c) not in visited and grid.get((r, c)) is not None and str(grid.get((r,c))).strip() != ""]
        if not row_nonvisited_nonempty:
            r += 1
            continue

        # Try detect table starting at this row
        tbl = detect_table_block(grid, visited, r, max_row, max_col)
        if tbl:
            sr, er, sc, ec = tbl
            headers, rows = rows_to_table_dicts(grid, sr, er, sc, ec, visited)
            items.append({
                "type": "table",
                "headers": headers,
                "rows": rows,
                "row": sr,
                "col": sc
            })
            r = er + 1
            continue

        # Otherwise, create text clusters from contiguous unvisited non-empty cells
        clusters = []
        cl = []
        for c in range(max_col):
            if (r, c) not in visited and grid.get((r, c)) is not None and str(grid.get((r,c))).strip() != "":
                cl.append(c)
            else:
                if cl:
                    clusters.append(cl)
                    cl = []
        if cl:
            clusters.append(cl)

        for cluster in clusters:
            text_pieces = []
            first_c = cluster[0]
            for c in cluster:
                v = grid.get((r, c))
                if v is not None and str(v).strip() != "":
                    text_pieces.append(str(v).strip())
                visited.add((r, c))
            combined = " ".join(text_pieces)
            items.append({
                "type": "text",
                "content": combined,
                "row": r,
                "col": first_c
            })
        r += 1

    # Append images that anchor beyond last row (if any) - avoid adding duplicates
    extra_images = [img for img in images_info if img["row"] >= max_row]
    for img in sorted(extra_images, key=lambda x: (x["row"], x["col"])):
        # ensure not already present (by path)
        if not any(it["type"] == "image" and it.get("path") == img["path"] for it in items):
            items.append({
                "type": "image",
                "path": img["path"],
                "description": img["desc"],
                "row": img["row"],
                "col": img["col"]
            })

    # Final sort by (row, col)
    items_sorted = sorted(items, key=lambda x: (x.get("row", 0), x.get("col", 0)))

    # Deduplicate while preserving order using deterministic keys
    seen = set()
    deduped = []
    for it in items_sorted:
        if it["type"] == "text":
            key = ("text", it["row"], it["col"], it["content"].strip())
        elif it["type"] == "image":
            key = ("image", it["row"], it["col"], os.path.basename(it["path"]))
        elif it["type"] == "table":
            # stringify headers + first few rows to form stable key
            headers_key = tuple(it.get("headers", []))
            rows_preview = tuple(tuple(str(v) for v in list(r.values())[:5]) for r in it.get("rows", [])[:3])
            key = ("table", it["row"], it["col"], headers_key, rows_preview)
        else:
            key = (it.get("type"), it.get("row"), it.get("col"))
        if key in seen:
            continue
        seen.add(key)
        deduped.append(it)

    return deduped

def extract_xlsx_ordered(filepath, get_image_desc_fn=get_image_description):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    res = {}
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        res[sheetname] = extract_ordered_sheet(ws, get_image_desc_fn)
    return res

def save_as_json_and_md(workbook_data, base_filename):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    json_path = os.path.join(OUTPUT_DIR, f"{base_filename}.json")
    md_path = os.path.join(OUTPUT_DIR, f"{base_filename}.md")

    with open(json_path, "w", encoding="utf-8") as jf:
        json.dump(workbook_data, jf, indent=2, ensure_ascii=False)

    with open(md_path, "w", encoding="utf-8") as mf:
        for sheet, items in workbook_data.items():
            mf.write(f"# Sheet: {sheet}\n\n")
            for item in items:
                if item["type"] == "text":
                    mf.write(f"{item['content']}\n\n")
                elif item["type"] == "table":
                    headers = item.get("headers", [])
                    rows = item.get("rows", [])
                    if headers:
                        mf.write("| " + " | ".join(headers) + " |\n")
                        mf.write("| " + " | ".join(["---"] * len(headers)) + " |\n")
                        for rdata in rows:
                            mf.write("| " + " | ".join(str(rdata.get(h, "")) for h in headers) + " |\n")
                        mf.write("\n")
                elif item["type"] == "image":
                    path = item.get("path")
                    desc = item.get("description", "")
                    rel = os.path.relpath(path, start=os.path.dirname(md_path))
                    mf.write(f"![{os.path.basename(path)}]({rel})\n\n")
                    mf.write(f"**Description:** {desc}\n\n")
            mf.write("\n---\n\n")
    return json_path, md_path

# ----------------- Entry point -----------------
if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python xlsx_ordered_extractor_no_dup_fixed.py <path_to_xlsx>")
        raise SystemExit(1)
    filepath = sys.argv[1]
    base = os.path.splitext(os.path.basename(filepath))[0]
    workbook_data = extract_xlsx_ordered(filepath)
    jpath, mpath = save_as_json_and_md(workbook_data, base)
    print("Saved:", jpath, mpath)
