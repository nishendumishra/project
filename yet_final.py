import openpyxl
from openpyxl.drawing.image import Image as XLImage
import json
import os
from PIL import Image
from io import BytesIO
import requests
import re

# Vision model endpoint and model name
VISION_MODEL_URI = "http://nip1gpu37.sdi.corp.bankofamerica.com:8000/v2/models/meta-llama_Llama-3.2-90B-Vision-Instruct/generate"
MODEL_NAME = "Llama-3.2-90B-Vision-Instruct"

# --------------------- Utility Functions ---------------------

def get_image_description(image_path):
    """Call the Llama Vision API to describe an image."""
    try:
        with open(image_path, "rb") as image_file:
            image_data = image_file.read()
            response = requests.post(
                VISION_MODEL_URI,
                headers={"Content-Type": "application/json"},
                json={
                    "model": MODEL_NAME,
                    "messages": [
                        {
                            "role": "user",
                            "content": "Describe this image. Also include any text present in the image."
                        }
                    ],
                    "images": [
                        {
                            "name": os.path.basename(image_path),
                            "data": image_data.decode("latin1")
                        }
                    ]
                }
            )
        response.raise_for_status()
        return response.json().get("message", {}).get("content", "No description available.")
    except Exception as e:
        return f"Error generating description: {e}"

def get_image_anchor(image):
    """Extract the anchor position (row, col) of an image in Excel."""
    anchor = getattr(image, "anchor", None)
    if not anchor:
        return None
    frm = getattr(anchor, "_from", None) or getattr(anchor, "from_", None) or getattr(anchor, "from", None)
    if frm:
        row, col = getattr(frm, "row", None), getattr(frm, "col", None)
        if row is not None and col is not None:
            return int(row), int(col)
    return None

def is_bullet(cell_value):
    """Detect if a cell starts with a bullet or list marker."""
    if cell_value is None:
        return False
    return bool(re.match(r'^[•\-\*\u2022]\s*', str(cell_value).strip()))

def is_probable_table(rows):
    """Check if a set of rows forms a genuine table."""
    if len(rows) < 2:
        return False

    non_empty_counts = []
    bullet_counts = []
    for row in rows:
        non_empty = sum(1 for c in row if c and str(c).strip() != "")
        bullets = sum(1 for c in row if is_bullet(c))
        non_empty_counts.append(non_empty)
        bullet_counts.append(bullets)

    avg_non_empty = sum(non_empty_counts) / len(non_empty_counts)
    avg_bullets = sum(bullet_counts) / len(bullet_counts)

    # Must have at least 2 filled columns across multiple rows
    if avg_non_empty >= 2:
        # But if majority are bullet-like rows, treat as plain text
        if avg_bullets / max(avg_non_empty, 1) > 0.3:
            return False
        return True
    return False

# --------------------- Core Extraction ---------------------

def extract_xlsx_content(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    workbook_data = {}

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        grid = {}
        max_row, max_col = ws.max_row, ws.max_column

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                grid[(r - 1, c - 1)] = ws.cell(row=r, column=c).value

        # Collect images
        images_info = []
        for idx, image in enumerate(ws._images):
            if isinstance(image, XLImage):
                try:
                    img_data = image._data()
                except Exception:
                    img_data = None
                if img_data:
                    img = Image.open(BytesIO(img_data))
                else:
                    img = Image.new("RGB", (100, 40), color=(255, 255, 255))

                img_name = f"{sheetname}_image_{idx + 1}.png"
                output_dir = "output_images"
                os.makedirs(output_dir, exist_ok=True)
                img_path = os.path.join(output_dir, img_name)
                img.save(img_path)

                anchor = get_image_anchor(image)
                row, col = anchor if anchor else (max_row, idx)
                description = get_image_description(img_path)

                images_info.append({
                    "type": "image",
                    "path": img_path,
                    "description": description,
                    "row": row,
                    "col": col
                })

        # Extract tables and text
        items = []
        r = 0
        while r < max_row:
            # Insert images anchored at this row (sorted by col)
            for img in sorted([i for i in images_info if i["row"] == r], key=lambda x: x["col"]):
                items.append(img)

            row_vals = [grid.get((r, c)) for c in range(max_col)]
            non_empty_cells = [c for c in range(max_col) if row_vals[c] and str(row_vals[c]).strip() != ""]

            if len(non_empty_cells) > 1:
                # Gather consecutive rows
                table_rows, start_r = [], r
                while r < max_row:
                    row_vals = [grid.get((r, c)) for c in range(max_col)]
                    if any(row_vals):
                        table_rows.append(row_vals)
                        r += 1
                    else:
                        break

                # Apply refined table check
                if is_probable_table(table_rows):
                    headers = [str(h) if h else f"Column_{i+1}" for i, h in enumerate(table_rows[0])]
                    table_dicts = [
                        {h: row[i] for i, h in enumerate(headers)} for row in table_rows[1:]
                    ]
                    items.append({
                        "type": "table",
                        "data": table_dicts,
                        "headers": headers,
                        "row": start_r,
                        "col": 0
                    })
                else:
                    # Treat as text instead of table
                    for row in table_rows:
                        text_parts = [str(cell).strip() for cell in row if cell and str(cell).strip()]
                        if text_parts:
                            text = " ".join(text_parts)
                            items.append({
                                "type": "text",
                                "content": text,
                                "row": r,
                                "col": 0
                            })
            else:
                text = next((str(cell).strip() for cell in row_vals if cell and str(cell).strip()), None)
                if text:
                    items.append({
                        "type": "text",
                        "content": text,
                        "row": r,
                        "col": 0
                    })
                r += 1

        # Add images anchored below last row
        for img in sorted([i for i in images_info if i["row"] >= max_row], key=lambda x: (x["row"], x["col"])):
            if not any(it["type"] == "image" and it.get("path") == img["path"] for it in items):
                items.append(img)

        items_sorted = sorted(items, key=lambda x: (x.get("row", 0), x.get("col", 0)))
        workbook_data[sheetname] = items_sorted

    return workbook_data

# --------------------- Output Writers ---------------------

def save_as_json_and_md(workbook_data, base_filename):
    os.makedirs("output", exist_ok=True)
    json_path = f"output/{base_filename}.json"
    md_path = f"output/{base_filename}.md"

    with open(json_path, "w", encoding="utf-8") as jf:
        json.dump(workbook_data, jf, indent=4, ensure_ascii=False)

    with open(md_path, "w", encoding="utf-8") as mf:
        for sheet, items in workbook_data.items():
            mf.write(f"# Sheet: {sheet}\n\n")
            for item in items:
                if item["type"] == "text":
                    mf.write(f"{item['content']}\n\n")
                elif item["type"] == "table":
                    headers = item["headers"]
                    table = item["data"]
                    if not headers or not table:
                        continue
                    mf.write("| " + " | ".join(headers) + " |\n")
                    mf.write("| " + " | ".join(["---"] * len(headers)) + " |\n")
                    for row in table:
                        mf.write("| " + " | ".join(str(row.get(h, "")) for h in headers) + " |\n")
                    mf.write("\n")
                elif item["type"] == "image":
                    rel = os.path.relpath(item['path'], start=os.path.dirname(md_path))
                    mf.write(f"![{os.path.basename(item['path'])}]({rel})\n")
                    mf.write(f"**Description:** {item['description']}\n\n")
            mf.write("\n---\n\n")

    return json_path, md_path

# --------------------- Main ---------------------

if __name__ == "__main__":
    filepath = "test_excel_with_image.xlsx"  # your input file
    base_filename = os.path.splitext(os.path.basename(filepath))[0]

    workbook_data = extract_xlsx_content(filepath)
    json_file, md_file = save_as_json_and_md(workbook_data, base_filename)
    print(f"✅ Extraction complete!\nJSON: {json_file}\nMarkdown: {md_file}")



import openpyxl
import json
import os
from PIL import Image
from io import BytesIO
import zipfile
import xml.etree.ElementTree as ET
import requests
import re

VISION_MODEL_URI = "http://nip1gpu37.sdi.corp.bankofamerica.com:8000/v2/models/meta-llama_Llama-3.2-90B-Vision-Instruct/generate"
MODEL_NAME = "Llama-3.2-90B-Vision-Instruct"


# --------------------- Vision Utility ---------------------

def get_image_description(image_path):
    """Describe image via Vision model (placeholder for now)."""
    try:
        with open(image_path, "rb") as f:
            image_data = f.read()
        response = requests.post(
            VISION_MODEL_URI,
            headers={"Content-Type": "application/json"},
            json={
                "model": MODEL_NAME,
                "messages": [{"role": "user", "content": "Describe this image and any text in it."}],
                "images": [{
                    "name": os.path.basename(image_path),
                    "data": image_data.decode("latin1")
                }]
            },
        )
        response.raise_for_status()
        return response.json().get("message", {}).get("content", "No description available.")
    except Exception as e:
        return f"Error generating description: {e}"


# --------------------- Text/Table Utilities ---------------------

def is_bullet(val):
    if val is None:
        return False
    return bool(re.match(r'^[•\-\*\u2022]\s*', str(val).strip()))


def is_probable_table(rows):
    if len(rows) < 2:
        return False
    non_empty_counts = []
    bullet_counts = []
    for row in rows:
        non_empty = sum(1 for c in row if c and str(c).strip())
        bullets = sum(1 for c in row if is_bullet(c))
        non_empty_counts.append(non_empty)
        bullet_counts.append(bullets)
    avg_non_empty = sum(non_empty_counts) / len(non_empty_counts)
    avg_bullets = sum(bullet_counts) / len(bullet_counts)
    return avg_non_empty >= 2 and avg_bullets / max(avg_non_empty, 1) < 0.3


# --------------------- Image Extraction via Zip ---------------------

def extract_all_images_from_zip(xlsx_path, output_dir):
    """Extract *all* images from xl/media and map them approximately to sheets."""
    os.makedirs(output_dir, exist_ok=True)
    images_info = []

    with zipfile.ZipFile(xlsx_path, "r") as zf:
        media_files = [f for f in zf.namelist() if f.startswith("xl/media/")]

        # Step 1: Save all images
        for idx, mf in enumerate(media_files, 1):
            data = zf.read(mf)
            img = Image.open(BytesIO(data))
            ext = os.path.splitext(mf)[1].lower() or ".png"
            img_path = os.path.join(output_dir, f"image_{idx}{ext}")
            img.save(img_path)
            images_info.append({
                "type": "image",
                "path": img_path,
                "sheet_guess": None,
                "row": idx,
                "col": 0
            })

        # Step 2: Map to sheets if possible (drawing relationships)
        drawing_map = {}  # drawing#.xml → [sheetname]
        sheet_drawing_map = {}

        # Read sheet rels
        for f in zf.namelist():
            if f.startswith("xl/worksheets/_rels/") and f.endswith(".rels"):
                xml = zf.read(f)
                tree = ET.fromstring(xml)
                for rel in tree.findall("{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"):
                    target = rel.attrib.get("Target", "")
                    if "drawing" in target:
                        sheet_name = os.path.basename(f).replace(".rels", "").replace("sheet", "Sheet ")
                        drawing_map[target.replace("../drawings/", "")] = sheet_name

        # Step 3: Link media → drawing → sheet
        for f in zf.namelist():
            if f.startswith("xl/drawings/_rels/") and f.endswith(".rels"):
                xml = zf.read(f)
                tree = ET.fromstring(xml)
                drawing_file = os.path.basename(f).replace(".rels", "")
                related_sheet = drawing_map.get(drawing_file)
                for rel in tree.findall("{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"):
                    target = rel.attrib.get("Target", "")
                    if "media" in target:
                        media_name = os.path.basename(target)
                        for img in images_info:
                            if os.path.basename(img["path"]).startswith(media_name.replace(".png", "").replace(".jpeg", "")):
                                img["sheet_guess"] = related_sheet or "Unknown"

    return images_info


# --------------------- Excel Cell/Text Extraction ---------------------

def extract_sheet_content(ws):
    """Extract text + tables from a single worksheet."""
    grid = {(r - 1, c - 1): ws.cell(row=r, column=c).value for r in range(1, ws.max_row + 1)
            for c in range(1, ws.max_column + 1)}

    items = []
    r, max_row, max_col = 0, ws.max_row, ws.max_column
    while r < max_row:
        row_vals = [grid.get((r, c)) for c in range(max_col)]
        non_empty_cells = [c for c in range(max_col) if row_vals[c] and str(row_vals[c]).strip()]

        if len(non_empty_cells) > 1:
            table_rows, start_r = [], r
            while r < max_row:
                row_vals = [grid.get((r, c)) for c in range(max_col)]
                if any(row_vals):
                    table_rows.append(row_vals)
                    r += 1
                else:
                    break
            if is_probable_table(table_rows):
                headers = [str(h) if h else f"Column_{i+1}" for i, h in enumerate(table_rows[0])]
                data = [{h: row[i] for i, h in enumerate(headers)} for row in table_rows[1:]]
                items.append({"type": "table", "headers": headers, "data": data, "row": start_r, "col": 0})
            else:
                for row in table_rows:
                    text = " ".join(str(cell).strip() for cell in row if cell and str(cell).strip())
                    if text:
                        items.append({"type": "text", "content": text, "row": r, "col": 0})
        else:
            text = next((str(cell).strip() for cell in row_vals if cell and str(cell).strip()), None)
            if text:
                items.append({"type": "text", "content": text, "row": r, "col": 0})
            r += 1

    return items


# --------------------- Main Extraction ---------------------

def extract_xlsx_content(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    workbook_data = {}

    # Text + table extraction
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        items = extract_sheet_content(ws)
        workbook_data[sheetname] = items

    # Image extraction (full coverage)
    all_images = extract_all_images_from_zip(filepath, "output_images")

    # Add vision description
    for img in all_images:
        img["description"] = get_image_description(img["path"])

    # Assign images to sheets (by guess)
    for img in all_images:
        sheet = img.get("sheet_guess") or wb.sheetnames[0]
        workbook_data.setdefault(sheet, []).append(img)

    return workbook_data


# --------------------- Output Writers ---------------------

def save_as_json_and_md(workbook_data, base_filename):
    os.makedirs("output", exist_ok=True)
    json_path = f"output/{base_filename}.json"
    md_path = f"output/{base_filename}.md"

    with open(json_path, "w", encoding="utf-8") as jf:
        json.dump(workbook_data, jf, indent=4, ensure_ascii=False)

    with open(md_path, "w", encoding="utf-8") as mf:
        for sheet, items in workbook_data.items():
            mf.write(f"# Sheet: {sheet}\n\n")
            for item in items:
                if item["type"] == "text":
                    mf.write(f"{item['content']}\n\n")
                elif item["type"] == "table":
                    headers = item["headers"]
                    table = item["data"]
                    if not headers or not table:
                        continue
                    mf.write("| " + " | ".join(headers) + " |\n")
                    mf.write("| " + " | ".join(["---"] * len(headers)) + " |\n")
                    for row in table:
                        mf.write("| " + " | ".join(str(row.get(h, "")) for h in headers) + " |\n")
                    mf.write("\n")
                elif item["type"] == "image":
                    rel = os.path.relpath(item["path"], start=os.path.dirname(md_path))
                    mf.write(f"![{os.path.basename(item['path'])}]({rel})\n")
                    mf.write(f"**Description:** {item['description']}\n\n")
            mf.write("\n---\n\n")

    return json_path, md_path


# --------------------- Main ---------------------

if __name__ == "__main__":
    filepath = "test_excel_with_image.xlsx"
    base_filename = os.path.splitext(os.path.basename(filepath))[0]

    workbook_data = extract_xlsx_content(filepath)
    json_file, md_file = save_as_json_and_md(workbook_data, base_filename)
    print(f"✅ Extraction complete!\nJSON: {json_file}\nMarkdown: {md_file}")
