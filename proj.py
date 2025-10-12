import openpyxl
from openpyxl.drawing.image import Image as XLImage
import json
import os
from PIL import Image
from io import BytesIO
import ollama  # Assuming ollama is the library to interact with LLaMA models

# ========== LLaMA Vision Model Integration (Stub) ==========
def get_image_description(image_path):
    try:
        response = ollama.chat(
            model='llama3.2-vision',
            messages=[{
                'role': 'user',
                'content': 'Describe this image. Also include any text present in the image.',
                'images': [image_path]
            }]
        )
        return response['message']['content']
    except Exception as e:
        return f"Error generating description: {e}"
    

# ===========================================================

def extract_xlsx_content(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    workbook_data = {}

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        sheet_data = {
            "PlainText": [],
            "Tables": [],
            "Images": []
        }

        # Extract text and detect tables
        raw_data = []
        for row in ws.iter_rows(values_only=True):
            raw_data.append(list(row))

        current_table = []
        for row in raw_data:
            if all(cell is None for cell in row):
                if current_table:
                    process_table(current_table, sheet_data)
                    current_table = []
                continue

            if sum(cell is not None for cell in row) > 1:
                current_table.append(row)
            else:
                if current_table:
                    process_table(current_table, sheet_data)
                    current_table = []
                text = next((cell for cell in row if cell is not None), None)
                if text:
                    sheet_data["PlainText"].append(str(text))

        if current_table:
            process_table(current_table, sheet_data)

        # Extract images and generate captions
        for image in ws._images:
            if isinstance(image, XLImage):
                img_data = image._data()
                img = Image.open(BytesIO(img_data))
                img_name = f"{sheetname}_image_{len(sheet_data['Images']) + 1}.png"
                output_dir = "output_images"
                os.makedirs(output_dir, exist_ok=True)
                img_path = os.path.join(output_dir, img_name)
                img.save(img_path)

                # Get image description using vision model
                description = get_image_description(img_path)

                sheet_data["Images"].append({
                    "path": img_path,
                    "description": description
                })

        workbook_data[sheetname] = sheet_data

    return workbook_data


def process_table(table_rows, sheet_data):
    headers = [str(h) if h else f"Column_{i+1}" for i, h in enumerate(table_rows[0])]
    table_dicts = []
    for row in table_rows[1:]:
        entry = {h: val for h, val in zip(headers, row)}
        table_dicts.append(entry)
    sheet_data["Tables"].append(table_dicts)


def save_as_json_and_md(workbook_data, base_filename):
    os.makedirs("output", exist_ok=True)
    json_path = f"output/{base_filename}.json"
    md_path = f"output/{base_filename}.md"

    # Save JSON
    with open(json_path, "w", encoding="utf-8") as jf:
        json.dump(workbook_data, jf, indent=4, ensure_ascii=False)

    # Save Markdown
    with open(md_path, "w", encoding="utf-8") as mf:
        for sheet, content in workbook_data.items():
            mf.write(f"# Sheet: {sheet}\n\n")

            if content["PlainText"]:
                mf.write("## Plain Text\n")
                for text in content["PlainText"]:
                    mf.write(f"- {text}\n")
                mf.write("\n")

            if content["Tables"]:
                mf.write("## Tables\n")
                for table in content["Tables"]:
                    if not table:
                        continue
                    headers = table[0].keys()
                    mf.write("| " + " | ".join(headers) + " |\n")
                    mf.write("| " + " | ".join(["---"] * len(headers)) + " |\n")
                    for row in table:
                        mf.write("| " + " | ".join(str(v) if v is not None else "" for v in row.values()) + " |\n")
                    mf.write("\n")

            if content["Images"]:
                mf.write("## Images\n")
                for img_info in content["Images"]:
                    mf.write(f"![{os.path.basename(img_info['path'])}]({img_info['path']})\n")
                    mf.write(f"**Description:** {img_info['description']}\n\n")

    return json_path, md_path


if __name__ == "__main__":
    filepath = "test_excel_with_image.xlsx"  # your input file
    base_filename = os.path.splitext(os.path.basename(filepath))[0]

    workbook_data = extract_xlsx_content(filepath)
    json_file, md_file = save_as_json_and_md(workbook_data, base_filename)
    print(f"✅ Extraction complete!\nJSON: {json_file}\nMarkdown: {md_file}")
