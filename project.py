import os
import json
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
from io import BytesIO

# --------- Configure ----------
OUTPUT_DIR = "output"
IMAGE_DIR = os.path.join(OUTPUT_DIR, "images")
os.makedirs(IMAGE_DIR, exist_ok=True)
# -----------------------------

def get_image_description(image_path, use_ollama=False):
    """
    Replace this placeholder with your real LLaMA/Ollama call.
    If use_ollama=True and ollama is installed and running, this will try to use it.
    """
    if use_ollama:
        try:
            import ollama
            resp = ollama.chat(
                model="llama3.2-vision",
                messages=[{
                    "role": "user",
                    "content": "Describe this image in one sentence. Include any visible text.",
                    "images": [image_path]
                }]
            )
            return resp["message"]["content"]
        except Exception as e:
            return f"[Error getting description from Ollama: {e}]"
    # default placeholder
    return f"Auto-generated description for {os.path.basename(image_path)}"

def expand_merged_cells(ws):
    """
    Return a dict mapping (r,c) 0-based -> value,
    where merged cell top-left value is propagated to all cells in the merged range.
    """
    cell_map = {}
    # Fill with direct values first
    max_row = ws.max_row
    max_col = ws.max_column
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            val = ws.cell(row=r, column=c).value
            cell_map[(r-1, c-1)] = val

    # Now propagate merged ranges
    for merged in ws.merged_cells.ranges:
        # merged is a MergedCellRange object, e.g. 'A1:C1'
        min_row, min_col, max_row, max_col = merged.min_row, merged.min_col, merged.max_row, merged.max_col
        top_left_val = ws.cell(row=min_row, column=min_col).value
        for r in range(min_row-1, max_row):
            for c in range(min_col-1, max_col):
                cell_map[(r, c)] = top_left_val
    return cell_map, ws.max_row, ws.max_column

def image_anchor_position(image, default_row, default_col):
    """
    Return (row_index_0based, col_index_0based) of the image anchor.
    Works with different anchor shapes used by openpyxl.
    If anchor not available, returns provided defaults.
    """
    anchor = getattr(image, 'anchor', None)
    # Many openpyxl anchors have attributes ._from or .from_
    try:
        if anchor is None:
            return default_row, default_col
        # prefer anchor._from (older/newer implementations)
        from_attr = getattr(anchor, '_from', None) or getattr(anchor, 'from_', None)
        if from_attr:
            row = getattr(from_attr, 'row', None)
            col = getattr(from_attr, 'col', None)
            if row is not None and col is not None:
                return int(row), int(col)
        # fallback: anchor has .row/.col
        row = getattr(anchor, 'row', None) or getattr(anchor, 'row_coord', None)
        col = getattr(anchor, 'col', None) or getattr(anchor, 'col_coord', None)
        if row is not None and col is not None:
            return int(row), int(col)
    except Exception:
        pass
    return default_row, default_col

def is_table_block(block_rows):
    """
    Heuristic: treat a block as table if:
      - it contains >=2 rows AND
      - at least one row in the block has >=2 non-empty cells
    This handles complex tables (including merged headers).
    """
    if len(block_rows) < 2:
        return False
    for row in block_rows:
        non_empty = sum(1 for v in row if v is not None and str(v).strip() != "")
        if non_empty >= 2:
            return True
    return False

def choose_header_row(block_rows):
    """
    Choose header row index inside block_rows.
    Heuristic: pick the first row with >=2 non-empty cells; if none, pick first row.
    Returns (header_index_in_block, headers_list)
    """
    for i, row in enumerate(block_rows):
        non_empty_cells = [str(v).strip() if v is not None else "" for v in row]
        if sum(1 for v in non_empty_cells if v != "") >= 2:
            headers = [h if h not in (None, "") else f"Column_{idx+1}" for idx, h in enumerate(row)]
            return i, headers
    # fallback
    headers = [h if h not in (None, "") else f"Column_{idx+1}" for idx, h in enumerate(block_rows[0])]
    return 0, headers

def rows_to_table_dicts(headers, data_rows):
    """
    Build list of dicts mapping headers -> row values.
    If data row shorter than headers, missing values -> None.
    """
    table = []
    for row in data_rows:
        entry = {}
        for i, h in enumerate(headers):
            val = row[i] if i < len(row) else None
            entry[str(h)] = val
        table.append(entry)
    return table

def extract_sheet_ordered(ws, get_desc_fn):
    """
    Returns a list of content items preserving order.
    Each item is a dict having at least: type, order, and content fields.
    type in {"text", "table", "image"}.
    order is numeric so final sort preserves sheet order.
    """
    cell_map, max_row, max_col = expand_merged_cells(ws)
    # Build full grid as list of rows
    grid = []
    for r in range(max_row):
        row = []
        for c in range(max_col):
            row.append(cell_map.get((r, c)))
        grid.append(row)

    # collect images with positions
    images_info = []
    for img_idx, image in enumerate(ws._images):
        # default fallback position: end of sheet
        default_row = max_row
        default_col = 0
        r0, c0 = image_anchor_position(image, default_row, default_col)
        # Save image binary
        try:
            img_data = image._data()
        except Exception:
            # some openpyxl versions store .ref or .path or .image - fallback to attribute
            try:
                img_data = image.ref  # unlikely
            except Exception:
                img_data = None
        # Save file
        img_name = f"{ws.title}_image_{img_idx+1}.png"
        img_path = os.path.join(IMAGE_DIR, img_name)
        try:
            if isinstance(img_data, (bytes, bytearray)):
                img = Image.open(BytesIO(img_data))
                img.save(img_path)
            else:
                # attempt to get from image.image if present
                inner = getattr(image, 'image', None)
                if inner is not None and hasattr(inner, 'fp'):
                    inner.fp.seek(0)
                    im = Image.open(inner.fp)
                    im.save(img_path)
                else:
                    # last resort: try to save from pillow Image object if image is already PIL
                    if hasattr(image, '_data'):
                        b = image._data()
                        im = Image.open(BytesIO(b))
                        im.save(img_path)
                    else:
                        # create placeholder image so pipeline doesn't fail
                        placeholder = Image.new("RGB", (100, 40), color=(200,200,200))
                        placeholder.save(img_path)
        except Exception:
            # create placeholder
            placeholder = Image.new("RGB", (100, 40), color=(200,200,200))
            placeholder.save(img_path)

        # description via model
        description = get_desc_fn(img_path)

        images_info.append({
            "type": "image",
            "path": img_path,
            "description": description,
            "anchor_row": int(r0),
            "anchor_col": int(c0),
            # compute order: row*10000 + col to preserve order within row
            "order": int(r0) * 10000 + int(c0)
        })

    # Build ordered content by scanning rows and injecting images where anchors match row
    content = []
    r = 0
    num_rows = len(grid)
    # Build quick lookup of images by anchor row
    images_by_row = {}
    for img in images_info:
        images_by_row.setdefault(img["anchor_row"], []).append(img)

    while r < num_rows:
        # Insert images anchored at this row (in order by column)
        if r in images_by_row:
            images_here = sorted(images_by_row[r], key=lambda x: x["anchor_col"])
            for img in images_here:
                content.append({
                    "type": "image",
                    "path": img["path"],
                    "description": img["description"],
                    "order": img["order"]
                })
        # If current row empty (all cells None/blank), skip and increment
        row_vals = grid[r]
        if all((v is None or str(v).strip() == "") for v in row_vals):
            r += 1
            continue

        # Otherwise gather a block until a blank row or an image anchor row (so images appear in right place)
        block_rows = []
        block_start = r
        while r < num_rows:
            # stop block if this row has an image anchored BEFORE its data should appear
            if r in images_by_row:
                # If image column is before where text would start we might want to break; but to keep simple,
                # we break block on image row so image appears separately at same row.
                break
            row_vals = grid[r]
            if all((v is None or str(v).strip() == "") for v in row_vals):
                break
            block_rows.append(row_vals)
            r += 1

        # classify block as table or text
        if is_table_block(block_rows):
            # choose header row inside block (index)
            header_idx_in_block, headers = choose_header_row(block_rows)
            # data rows = block_rows after header row
            data_rows = block_rows[header_idx_in_block+1:]
            if not data_rows:
                # if no data rows treat header row as one-row table (still produce empty table)
                data_rows = []
            table_dicts = rows_to_table_dicts(headers, data_rows)
            content.append({
                "type": "table",
                "data": table_dicts,
                "headers": headers,
                "order": block_start * 10000
            })
        else:
            # treat each non-empty cell in the block as text content, preserve row order
            for i, brow in enumerate(block_rows):
                text = next((str(c) for c in brow if c is not None and str(c).strip() != ""), None)
                if text:
                    content.append({
                        "type": "text",
                        "content": text,
                        "order": (block_start + i) * 10000
                    })
        # continue loop (r already moved to after block)
    # It's possible some images anchor beyond last data row - append them
    extra_images = [img for img in images_info if img["anchor_row"] >= num_rows]
    for img in sorted(extra_images, key=lambda x: x["order"]):
        content.append({
            "type": "image",
            "path": img["path"],
            "description": img["description"],
            "order": img["order"]
        })

    # finally sort content by order and return
    content_sorted = sorted(content, key=lambda x: x["order"])
    return content_sorted

def extract_xlsx_ordered(filepath, get_desc_fn=get_image_description):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    workbook_data = {}
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        sheet_ordered = extract_sheet_ordered(ws, get_desc_fn)
        workbook_data[sheetname] = sheet_ordered
    return workbook_data

def save_as_json_and_md(workbook_data, base_filename):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    json_path = os.path.join(OUTPUT_DIR, f"{base_filename}.json")
    md_path = os.path.join(OUTPUT_DIR, f"{base_filename}.md")

    # Save JSON
    with open(json_path, "w", encoding="utf-8") as jf:
        json.dump(workbook_data, jf, indent=2, ensure_ascii=False)

    # Save MD
    with open(md_path, "w", encoding="utf-8") as mf:
        for sheet, items in workbook_data.items():
            mf.write(f"# Sheet: {sheet}\n\n")
            for item in items:
                if item["type"] == "text":
                    mf.write(f"- {item['content']}\n\n")
                elif item["type"] == "table":
                    headers = item.get("headers", [])
                    rows = item.get("data", [])
                    if headers:
                        mf.write("| " + " | ".join(map(str, headers)) + " |\n")
                        mf.write("| " + " | ".join(["---"] * len(headers)) + " |\n")
                        for r in rows:
                            mf.write("| " + " | ".join(str(r.get(h, "")) if r.get(h, "") is not None else "" for h in headers) + " |\n")
                        mf.write("\n")
                elif item["type"] == "image":
                    path = item.get("path")
                    desc = item.get("description", "")
                    mf.write(f"![{os.path.basename(path)}]({path})\n\n")
                    mf.write(f"**Description:** {desc}\n\n")
            mf.write("\n---\n\n")
    return json_path, md_path

# ---------------------- Usage ----------------------
if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python advanced_xlsx_ordered_extractor.py <path_to_xlsx> [--ollama]")
        sys.exit(1)
    filepath = sys.argv[1]
    use_ollama_flag = ("--ollama" in sys.argv)
    def desc_fn(p): return get_image_description(p, use_ollama=use_ollama_flag)

    base = os.path.splitext(os.path.basename(filepath))[0]
    workbook_data = extract_xlsx_ordered(filepath, get_desc_fn=desc_fn)
    jpath, mpath = save_as_json_and_md(workbook_data, base)
    print("Saved:", jpath, mpath)

import openpyxl
from openpyxl.drawing.image import Image as XLImage
import json
import os
from PIL import Image
from io import BytesIO
import requests


# -------------------- LLaMA Vision Description --------------------

VISION_MODEL_URI = "http://nip1gpu37.sdl.corp.bankofamerica.com:8000/v2/models/meta-llama_Llama-3.2-90B-Vision-Instruct/generate"
MODEL_NAME = "llama-3.2-90b-vision-instruct"

def get_image_description(image_path):
    """Get descriptive caption using LLaMA Vision model."""
    try:
        with open(image_path, "rb") as image_file:
            response = requests.post(
                VISION_MODEL_URI,
                headers={"Content-Type": "application/json"},
                json={
                    "model": MODEL_NAME,
                    "messages": [
                        {"role": "user", "content": "Describe this image. Include any text present in the image."}
                    ],
                    "images": [
                        {
                            "name": os.path.basename(image_path),
                            "data": image_file.read().decode("latin1"),
                        }
                    ],
                },
            )
        response.raise_for_status()
        return response.json().get("message", "No description available.")
    except Exception as e:
        return f"Error generating description: {e}"


# -------------------- Excel Extraction --------------------

def extract_xlsx_content(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    workbook_data = {}

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        sheet_items = []

        # 1️⃣ Extract all images and record their positions
        images = []
        for img_idx, image in enumerate(ws._images):
            if not isinstance(image, XLImage):
                continue

            img = Image.open(BytesIO(image._data()))
            os.makedirs("output/images", exist_ok=True)
            img_path = f"output/images/{sheetname}_img{img_idx+1}.png"
            img.save(img_path)

            # Get anchor position
            anchor = getattr(image, "anchor", None)
            if hasattr(anchor, "from_"):
                row, col = anchor.from_.row + 1, anchor.from_.col + 1
            else:
                row, col = 9999, img_idx  # fallback

            desc = get_image_description(img_path)
            images.append({
                "type": "image",
                "path": img_path,
                "description": desc,
                "row": row,
                "col": col,
            })

        # 2️⃣ Parse cells row by row (to maintain order)
        max_col = ws.max_column
        current_table = []
        table_start_row = None

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            row_values = [cell for cell in row]

            # Check if row has multiple filled cells → likely part of a table
            non_empty = [c for c in row_values if c not in (None, "")]
            if len(non_empty) > 1:
                if not current_table:
                    table_start_row = row_idx
                current_table.append(row_values)
            else:
                # If a table was ongoing, end it and append
                if current_table:
                    sheet_items.append({
                        "type": "table",
                        "data": convert_table_to_dict(current_table),
                        "row": table_start_row,
                        "col": 1,
                    })
                    current_table = []

                # If this row has a single cell of text
                text_content = non_empty[0] if non_empty else None
                if text_content:
                    sheet_items.append({
                        "type": "text",
                        "content": str(text_content).strip(),
                        "row": row_idx,
                        "col": 1,
                    })

        # Handle last table if file ends with table
        if current_table:
            sheet_items.append({
                "type": "table",
                "data": convert_table_to_dict(current_table),
                "row": table_start_row,
                "col": 1,
            })

        # 3️⃣ Merge image items in their order
        sheet_items.extend(images)
        # Sort everything by visual position
        sheet_items.sort(key=lambda x: (x["row"], x["col"]))

        workbook_data[sheetname] = sheet_items

    return workbook_data


def convert_table_to_dict(table_rows):
    """Convert a list of table rows into key-value dict list."""
    if not table_rows:
        return []

    headers = [
        str(h).strip() if h not in (None, "") else f"Column{i+1}"
        for i, h in enumerate(table_rows[0])
    ]
    table_dicts = []

    for row in table_rows[1:]:
        entry = {}
        for i, h in enumerate(headers):
            entry[h] = row[i] if i < len(row) and row[i] is not None else ""
        table_dicts.append(entry)

    return table_dicts


# -------------------- Save JSON + Markdown --------------------

def save_as_json_and_md(workbook_data, base_filename):
    os.makedirs("output", exist_ok=True)
    json_path = f"output/{base_filename}.json"
    md_path = f"output/{base_filename}.md"

    with open(json_path, "w", encoding="utf-8") as jf:
        json.dump(workbook_data, jf, indent=4, ensure_ascii=False)

    with open(md_path, "w", encoding="utf-8") as mf:
        for sheet, items in workbook_data.items():
            mf.write(f"# Sheet: {sheet}\n\n")
            for item in items:
                if item["type"] == "text":
                    mf.write(f"{item['content']}\n\n")
                elif item["type"] == "table":
                    if not item["data"]:
                        continue
                    headers = list(item["data"][0].keys())
                    mf.write("| " + " | ".join(headers) + " |\n")
                    mf.write("| " + " | ".join(["---"] * len(headers)) + " |\n")
                    for row in item["data"]:
                        mf.write("| " + " | ".join(str(row[h]) for h in headers) + " |\n")
                    mf.write("\n")
                elif item["type"] == "image":
                    mf.write(f"![Image]({item['path']})\n")
                    mf.write(f"**Description:** {item['description']}\n\n")

    return json_path, md_path


# -------------------- Main --------------------

if __name__ == "__main__":
    filepath = "2024b Online Assessment Package - Vendor, LTD..xlsx"  # your file
    base_filename = os.path.splitext(os.path.basename(filepath))[0]
    workbook_data = extract_xlsx_content(filepath)
    json_file, md_file = save_as_json_and_md(workbook_data, base_filename)
    print(f"✅ Extraction complete.\nJSON → {json_file}\nMarkdown → {md_file}")



import os
import json
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
from io import BytesIO

# ---------- CONFIG ----------
OUTPUT_DIR = "output"
IMAGE_DIR = os.path.join(OUTPUT_DIR, "images")
os.makedirs(IMAGE_DIR, exist_ok=True)
# ----------------------------

def get_image_description(image_path):
    """
    Placeholder: replace with Ollama or your HTTP LLaMA Vision API call.
    Keep it robust to exceptions and return a string.
    """
    try:
        # Example simple caption; replace with real call
        return f"Description for {os.path.basename(image_path)}"
    except Exception as e:
        return f"Error describing image: {e}"

def expand_merged_cells(ws):
    """
    Return a dict mapping (r,c) 0-based -> value and the actual grid size (rows,cols).
    Propagates merged cell top-left values to all cells in the merged range.
    """
    max_row = ws.max_row
    max_col = ws.max_column
    cell_map = {}
    # fill base values
    for r in range(1, max_row+1):
        for c in range(1, max_col+1):
            cell_map[(r-1, c-1)] = ws.cell(row=r, column=c).value

    # propagate merged ranges
    for merged in ws.merged_cells.ranges:
        min_row, min_col, maxr, maxc = merged.min_row, merged.min_col, merged.max_row, merged.max_col
        top_val = ws.cell(row=min_row, column=min_col).value
        for rr in range(min_row-1, maxr):
            for cc in range(min_col-1, maxc):
                cell_map[(rr,cc)] = top_val

    return cell_map, max_row, max_col

def image_anchor_pos(image):
    """
    Try to get 0-based (row, col) anchor for an openpyxl image.
    If not available, returns None.
    """
    anchor = getattr(image, "anchor", None)
    if not anchor:
        return None
    # try multiple possible attribute names robustly
    frm = getattr(anchor, "_from", None) or getattr(anchor, "from_", None) or getattr(anchor, "from", None)
    if frm:
        row = getattr(frm, "row", None)
        col = getattr(frm, "col", None)
        if row is not None and col is not None:
            return int(row), int(col)
    # fallback: try 'row'/'col' directly
    row = getattr(anchor, "row", None)
    col = getattr(anchor, "col", None)
    if row is not None and col is not None:
        return int(row), int(col)
    return None

def save_image_from_openpyxl(image_obj, sheetname, idx):
    """
    Save image bytes from openpyxl image object to IMAGE_DIR and return path.
    """
    # Try various ways to fetch binary image data safely
    img_path = os.path.join(IMAGE_DIR, f"{sheetname}_image_{idx+1}.png")
    try:
        data = None
        # Many openpyxl Image objects have ._data() that returns bytes
        if hasattr(image_obj, "_data"):
            data = image_obj._data()
        # Some have .ref or .image - attempt to access .image if it's a PIL image
        if data is None and hasattr(image_obj, "image"):
            inner = image_obj.image
            try:
                if hasattr(inner, "fp"):
                    inner.fp.seek(0)
                    im = PILImage.open(inner.fp)
                    im.save(img_path)
                    return img_path
                elif isinstance(inner, PILImage.Image):
                    inner.save(img_path)
                    return img_path
            except Exception:
                pass
        if data:
            im = PILImage.open(BytesIO(data))
            im.save(img_path)
            return img_path
    except Exception:
        pass

    # Fallback placeholder (so pipeline does not crash)
    placeholder = PILImage.new("RGB", (200,80), color=(230,230,230))
    placeholder.save(img_path)
    return img_path

def detect_table_block(grid, visited, start_r, max_rows, max_cols):
    """
    Try to detect a table starting at row start_r.
    - A table must have at least two consecutive rows where each row has >=2 unvisited non-empty cells.
    - Returns (start_row, end_row, start_col, end_col) in 0-based coordinates, or None if no table.
    """
    # scan ahead up to some reasonable number of rows to find consecutive rows with >=2 non-empty unvisited cells
    consecutive = 0
    candidate_rows = []
    r = start_r
    while r < max_rows and len(candidate_rows) < 50:  # limit lookahead
        row = [grid.get((r,c)) for c in range(max_cols)]
        # count unvisited non-empty cells
        non_empty_indices = [c for c in range(max_cols) if (r,c) not in visited and row[c] is not None and str(row[c]).strip() != ""]
        if len(non_empty_indices) >= 2:
            candidate_rows.append(non_empty_indices)
            consecutive += 1
            r += 1
            # require at least 2 consecutive rows of this form
            if consecutive >= 2:
                break
        else:
            break

    if consecutive < 2:
        return None

    # Determine column span as min..max of indices that commonly appear across candidate rows
    all_indices = [idx for sub in candidate_rows for idx in sub]
    if not all_indices:
        return None
    start_col = min(all_indices)
    end_col = max(all_indices)
    # Expand end_col/backwards a little if neighboring columns have values across rows
    # Return table bounding box
    end_row = start_r + len(candidate_rows)  # exclusive end in terms of next row index
    return (start_r, end_row-1, start_col, end_col)

def extract_ordered_sheet(ws, get_image_desc_fn):
    """
    Returns a list of items (in order) for this worksheet.
    Each item: {"type": "text"|"table"|"image", ... , "row":r, "col":c}
    row/col are 0-based positions used for ordering.
    """
    grid, max_row, max_col = expand_merged_cells(ws)
    visited = set()  # cells marked as consumed (r,c)
    items = []

    # Preprocess images with anchors and saved paths
    images_info = []
    for i, img in enumerate(ws._images):
        if not isinstance(img, XLImage):
            continue
        pos = image_anchor_pos(img)
        img_path = save_image_from_openpyxl(img, ws.title, i)
        desc = get_image_desc_fn(img_path)
        if pos is None:
            # anchor beyond sheet -> put after last row
            images_info.append({"row": max_row, "col": 0, "path": img_path, "desc": desc})
        else:
            images_info.append({"row": pos[0], "col": pos[1], "path": img_path, "desc": desc})

    # Build fast lookup of images by row
    images_by_row = {}
    for imginfo in images_info:
        images_by_row.setdefault(imginfo["row"], []).append(imginfo)

    r = 0
    while r < max_row:
        # 1) Insert any images anchored at this row, sorted by column
        if r in images_by_row:
            imgs_here = sorted(images_by_row[r], key=lambda x: x["col"])
            for imginfo in imgs_here:
                items.append({
                    "type": "image",
                    "path": imginfo["path"],
                    "description": imginfo["desc"],
                    "row": imginfo["row"],
                    "col": imginfo["col"]
                })

        # 2) If entire row is empty or consumed, skip
        row_vals = [grid.get((r,c)) for c in range(max_col)]
        row_unvisited_nonempty = [c for c in range(max_col) if (r,c) not in visited and row_vals[c] is not None and str(row_vals[c]).strip() != ""]
        if not row_unvisited_nonempty:
            r += 1
            continue

        # 3) Try detect a table starting at this row
        tbl_bbox = detect_table_block(grid, visited, r, max_row, max_col)
        if tbl_bbox:
            sr, er, sc, ec = tbl_bbox
            # Extract table rows from sr..er (inclusive), columns sc..ec (inclusive)
            # Use first row as header
            header_row = [grid.get((sr, c)) for c in range(sc, ec+1)]
            headers = []
            for i, h in enumerate(header_row):
                if h is None or str(h).strip() == "":
                    headers.append(f"Column_{i+1}")
                else:
                    headers.append(str(h))
            data_rows = []
            for rr in range(sr+1, er+1):
                row_dict = {}
                any_nonempty = False
                for ci, c in enumerate(range(sc, ec+1)):
                    val = grid.get((rr,c))
                    # mark visited cell
                    visited.add((rr,c))
                    if val is not None and str(val).strip() != "":
                        any_nonempty = True
                        row_dict[headers[ci]] = val
                    else:
                        row_dict[headers[ci]] = ""
                if any_nonempty:
                    data_rows.append(row_dict)
            # mark header cells visited too so text isn't re-read
            for c in range(sc, ec+1):
                visited.add((sr,c))

            items.append({
                "type": "table",
                "headers": headers,
                "rows": data_rows,
                "row": sr,
                "col": sc
            })
            # advance r to row after the table
            r = er + 1
            continue

        # 4) Otherwise treat cell(s) in this row as text blocks.
        # We'll group contiguous clusters of unvisited non-empty cells in row as one text item
        clusters = []
        cluster = []
        for c in range(max_col):
            if (r,c) not in visited and grid.get((r,c)) is not None and str(grid.get((r,c))).strip() != "":
                cluster.append(c)
            else:
                if cluster:
                    clusters.append(cluster)
                    cluster = []
        if cluster:
            clusters.append(cluster)
        for cl in clusters:
            # pick the left-most non-empty cell in cluster as representative text
            first_c = cl[0]
            text_pieces = []
            # Pull text from cluster cells (left->right) but join them with spaces
            for c in cl:
                val = grid.get((r,c))
                if val is not None and str(val).strip() != "":
                    text_pieces.append(str(val).strip())
                visited.add((r,c))  # mark as consumed
            combined = " ".join(text_pieces)
            items.append({
                "type": "text",
                "content": combined,
                "row": r,
                "col": first_c
            })
        r += 1

    # After scanning rows, append any images anchored beyond last row
    extra_images = [img for img in images_info if img["row"] >= max_row]
    for img in sorted(extra_images, key=lambda x: (x["row"], x["col"])):
        items.append({
            "type": "image",
            "path": img["path"],
            "description": img["desc"],
            "row": img["row"],
            "col": img["col"]
        })

    # Final sort by (row, col) to be safe
    items_sorted = sorted(items, key=lambda x: (x["row"], x["col"]))
    return items_sorted

def extract_xlsx_ordered(filepath, get_image_desc_fn=get_image_description):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    res = {}
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        res[sheetname] = extract_ordered_sheet(ws := ws, get_image_desc_fn=get_image_desc_fn)
    return res

def save_as_json_and_md(workbook_data, base_filename):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    json_path = os.path.join(OUTPUT_DIR, f"{base_filename}.json")
    md_path = os.path.join(OUTPUT_DIR, f"{base_filename}.md")

    with open(json_path, "w", encoding="utf-8") as jf:
        json.dump(workbook_data, jf, indent=2, ensure_ascii=False)

    with open(md_path, "w", encoding="utf-8") as mf:
        for sheet, items in workbook_data.items():
            mf.write(f"# Sheet: {sheet}\n\n")
            for item in items:
                if item["type"] == "text":
                    mf.write(f"{item['content']}\n\n")
                elif item["type"] == "table":
                    headers = item.get("headers", [])
                    rows = item.get("rows", [])
                    if headers:
                        mf.write("| " + " | ".join(headers) + " |\n")
                        mf.write("| " + " | ".join(["---"] * len(headers)) + " |\n")
                        for rdata in rows:
                            mf.write("| " + " | ".join(str(rdata.get(h, "")) for h in headers) + " |\n")
                        mf.write("\n")
                elif item["type"] == "image":
                    path = item.get("path")
                    desc = item.get("description", "")
                    # Use relative path so markdown renders locally
                    rel = os.path.relpath(path, start=os.path.dirname(md_path))
                    mf.write(f"![{os.path.basename(path)}]({rel})\n\n")
                    mf.write(f"**Description:** {desc}\n\n")
            mf.write("\n---\n\n")
    return json_path, md_path

# ----------------- Entry point -----------------
if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python xlsx_ordered_extractor_no_dup.py <path_to_xlsx>")
        raise SystemExit(1)
    filepath = sys.argv[1]
    base = os.path.splitext(os.path.basename(filepath))[0]
    workbook_data = extract_xlsx_ordered(filepath)
    jpath, mpath = save_as_json_and_md(workbook_data, base)
    print("Saved:", jpath, mpath)
